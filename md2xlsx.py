#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  mk2pyxl.py
#  
#  Copyright (c) 2018 Cilyan Olowen <gaknar@gmail.com>
#  
#  Permission is hereby granted, free of charge, to any person obtaining a copy
#  of this software and associated documentation files (the "Software"), to deal
#  in the Software without restriction, including without limitation the rights
#  to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
#  copies of the Software, and to permit persons to whom the Software is
#  furnished to do so, subject to the following conditions:
#  
#  The above copyright notice and this permission notice shall be included in all
#  copies or substantial portions of the Software.
#  
#  THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#  IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#  FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
#  AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
#  LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
#  OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
#  SOFTWARE.

import itertools
from collections import namedtuple

from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.colors import Color
from openpyxl.utils.indexed_list import IndexedList

import mistune

FT_HEADER1 = Font(name="Calibri", sz=26, family=2, b=False, i=False, color=Color(theme=1), scheme="minor")
FT_HEADER2 = Font(name="Calibri", sz=22, family=2, b=False, i=False, color=Color(theme=1), scheme="minor")
FT_HEADER3 = Font(name="Calibri", sz=18, family=2, b=False, i=False, color=Color(theme=1), scheme="minor")
FT_HEADER4 = Font(name="Calibri", sz=16, family=2, b=False, i=False, color=Color(theme=1), scheme="minor")
FT_HEADER5 = Font(name="Calibri", sz=14, family=2, b=False, i=False, color=Color(theme=1), scheme="minor")
FT_HEADER6 = Font(name="Calibri", sz=12, family=2, b=False, i=False, color=Color(theme=1), scheme="minor")

FILL_QUOTE = PatternFill(fill_type="solid", start_color='FFA09BBB', end_color='FFA09BBB')
FILL_HRULE = PatternFill(fill_type="solid", start_color='FF000000', end_color='FF000000')

class ExcelCollectorRenderer:
    """
        This class renders a BlockCollector into an OOXML worksheet.
        
        Each block item (paragraph, headline, code...) is rendered in a
        separate cell, while inline elements (including newlines) are
        placed into cells as rich text.
    """
    FONTS = {
        "h1": FT_HEADER1,
        "h2": FT_HEADER2,
        "h3": FT_HEADER3,
        "h4": FT_HEADER4,
        "h5": FT_HEADER5,
        "h6": FT_HEADER6
    }
    
    def __init__(self, worksheet, row = 1, col = 1):
        self.ws = worksheet
        self.row = itertools.count(row)
        self.col = col
    
    def render(self, block_collector):
        """
            Render the collector and links into the current worksheet at
            current position.
        """
        for row, block in zip(self.row, block_collector):
            cell = self.ws.cell(row=row, column=self.col)
            self.do_cell(cell, block)
    
    def do_cell(self, cell, block):
        """
            Fill and style ``cell`` according to ``cell_properties``.
        """
        self.prepare_cell(cell, block.content, block.styles)
        self.fill_content(cell, block.content)
        font, alignment, fill = self.create_style(block.styles)
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if fill is not None:
            cell.fill = fill
    
    def prepare_cell(self, cell, content, styles):
        """
            Prepare cell when styles influences content and vice-versa.
        """
        cell.rich_text = ""
        if "list_item" in styles:
            cell.rich_text += "• "
    
    def create_style(self, styles):
        """
            Return a Font, Alignment and PatterFill to be applied to cells
            that shall display the given list of styles.
        """
        font = None
        alignment = None
        fill = None
        for style in styles:
            if style in self.FONTS:
                font = self.FONTS[style]
            elif style in {"list", "link"}:
                if alignment is None:
                    alignment = Alignment()
                alignment.indent += 2
            elif style == "quote":
                if alignment is None:
                    alignment = Alignment()
                alignment.indent += 2
                if fill is None:
                    fill = FILL_QUOTE
        return font, alignment, fill
    
    def fill_content(self, cell, content):
        """
            Place the content in the cell as rich text
        """
        for element in content:
            font = self.create_font(element.markers)
            cell.rich_text += (element.text, font)
    
    def create_font(self, markers):
        """
            Returns a font suitable for rich text that has been tagged with the
            set of ``markers``.
        """
        if len(markers) == 0:
            return None
        return Font(
            name="Consolas" if "c" in markers else "Calibri",
            sz=11,
            family=3 if "c" in markers else 2,
            b=True if "b" in markers else False,
            i=True if "i" in markers else False,
            color= Color(rgb="FF382DA0") if "l" in markers else Color(theme=1),
            u='single' if "l" in markers else None,
            strike=True if "s" in markers else None
        )

BlockElement = namedtuple("BlockElement", ["content", "styles"])

class BlockCollector:
    """
        Collects block type elements
    """
    def __init__(self, content = None):
        self.elements = []
        if content is not None:
            self.append(content)
    def append(self, content):
        """
            Append a new block that will hold ``content``
        """
        styles = []
        self.elements.append(BlockElement(content, styles))
    def __iadd__(self, other):
        self.elements += other.elements
        return self
    def __iter__(self):
        return iter(self.elements)
    def nest(self, style):
        """
            Add a new style to all blocks in the collector, creating a
            "nesting" effect.
        """
        for element in self.elements:
            element.styles.insert(0, style)
        return self
    def __repr__(self):
        return "<{} {!r}>".format(self.__class__.__name__, self.elements)

InlineElement = namedtuple("InlineElement", ["text", "markers"])

class InlineCollector:
    def __init__(self, text = None):
        self.elements = []
        if text is not None:
            self.append(text)
    def append(self, text):
        """
            Append a new inline element holding ``text``
        """
        markers = set()
        self.elements.append(InlineElement(text, markers))
        return self
    def __iadd__(self, bucket):
        self.elements += bucket.elements
        return self
    def __iter__(self):
        return iter(self.elements)
    def apply_marker(self, marker):
        """
            Apply a new style marker to all elements of the collector
        """
        for element in self.elements:
            element.markers.add(marker)
        return self
    def __repr__(self):
        return "<{} {!r}>".format(self.__class__.__name__, self.elements)

class Collector:
    """
        The Collector is the object returned by the markdown processor when
        parsing of the tagged text is finished.
        
        The :py:func:`render` function can then be used to place the result
        of the collection inside a Worksheet.
    """
    # The Markdown processor of mistune works in a specific way. At each
    # start of a block or inline markup, it requests a "placeholder" to the
    # renderer. The objects returned afterwards by the renderer are appended
    # to this placeholder. Obviously, the concept has strings in mind.
    # But when working with OOXML sheets, we want to keep control of what
    # happens because appending to a cell or a run of rich text isn't as
    # simple as string concatenation.
    # We fake the markdown processor by sending the Collector as our
    # placeholder, which will act as a spy and collect everything sent back
    # chunk by chunk by the renderer using an overloading of the += operator.
    def __init__(self, renderer):
        self.renderer = renderer
        self.inline = None
        self.block = None
    
    def __iadd__(self, other):
        # Overload of the += operator. This is used by the markdown processor
        # to combine chunks of formatted text together. For our OOXML use
        # case, we will need to separate blocks which will end up as cells
        # and inlines which are pieces of a text inside a cell.
        # It is expected that a Collector contains only blocks or only inlines.
        if isinstance(other, BlockCollector):
            self._add_block(other)
        elif isinstance(other, InlineCollector):
            self._add_inline(other)
        elif isinstance(other, Collector):
            # Sometimes, the markdown processor tries to combine chunks without
            # asking the renderer for a transformation. When that happens,
            # our Collector receives another Collector. We want to merge
            # only Collectors of the same kind.
            if other.block is not None:
                self += other.block
                if other.inline:
                    # If there exists a trailing inline content along with
                    # block data, convert to block
                    self += BlockCollector(other.inline)
            else:
                self += other.inline
        elif other is None:
            # This handling is intended for unsupported markup
            pass
        else:
            raise RuntimeError(
                "Unexpected type {} appended to Collector: {!r}".format(
                    other.__class__.__name__,
                    other
                )
            )
        return self
    
    def _add_block(self, block_collector):
        # Append a block collector to the current collector
        if self.block is None:
            if self.inline is not None:
                # This special case happens when a list is nested in a
                # list_item
                self.block = BlockCollector(self.inline)
                self.inline = None
            else:
                self.block = BlockCollector()
        self.block += block_collector
    
    def _add_inline(self, inline_collector):
        # Append an inline collector to the current collector
        if self.inline is None:
            # There is a special case where block is also not None, when
            # a nested list is followed by a chunk of content targetted at the
            # parent list_item
            self.inline = inline_collector
        else:
            self.inline += inline_collector
    
    def __repr__(self):
        return "<{} {!r}, {!r}>".format(
            self.__class__.__name__,
            self.inline,
            self.block
        )
    
    def render(
            self,
            worksheet,
            row = 1, col = 1,
            renderer_class = ExcelCollectorRenderer
        ):
        """
            Renders the result of parsing into a worksheet. The output will
            be placed in column `col` starting at row `row`.
        """
        renderer = renderer_class(worksheet, row, col)
        if self.block is None:
            # If we collected only a sequence of inline elements, wrap it
            # into a cell element.
            self.block = BlockCollector(self.inline)
            self.inline = None
        # Allow the markdown renderer to send us last minute content
        self += self.renderer.terminate()
        # Call the excel renderer to do the heavy lifting
        renderer.render(self.block)


class ExcelRenderer(mistune.Renderer):
    """
        Renders a Markdown formatted text to an OOXML worksheet.
        Each block item (paragraph, headline, code...) is rendered in a
        separate cell, while inline elements (including newlines) are
        placed into cells as rich text.
        
        Some features of Markdown are not supported:
        - Tables
        - Images
        - HTML tags
        - Footnotes (for the moment)
        - List item continuation after a nested list (is considered as a new
          list item)
        
        Links are output in text as footnotes and target is placed at the end,
        after a blank cell.
        
        When using this renderer, the markdown processor will output an
        instance of :py:class:`Collector`. Use its :py:func:`~Collector.render`
        function to get the result in a worksheet.
        
        .. sourcecode:: python
        
            from openpyxl import Workbook
            
            renderer = ExcelRenderer()
            markdown = mistune.Markdown(renderer=renderer)
            collector = markdown(text)
            
            wb = Workbook()
            ws = wb.active
            collector.render(ws)
            wb.save(<output path>)
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.links = IndexedList()
        
    def placeholder(self):
        return Collector(self)
    
    # ==== Block elements ====
    
    def _get_block(self, bucket):
        if bucket.inline is not None:
            return BlockCollector(bucket.inline)
        else:
            return bucket.block
    
    def block_code(self, code, language=None):
        #print("block_code: ", code, language)
        return self._get_block(code).nest("code")
    
    def block_quote(self, text):
        #print("block_quote: ", text)
        return self._get_block(text).nest("quote")
    
    def block_html(self, html):
        #print("block_html: ", html)
        return None
    
    def header(self, text, level, raw=None):
        #print("header: ", text, level, raw)
        return self._get_block(text).nest("h"+str(level))
    
    def hrule(self):
        #print("hrule: ")
        return BlockCollector(InlineCollector("")).nest("hrule")
    
    def list(self, body, ordered=True):
        #print("list: ", body, ordered)
        return self._get_block(body).nest("list")
    
    def list_item(self, text):
        #print("list_item: ", text)
        return self._get_block(text).nest("list_item")
    
    def paragraph(self, text):
        #print("paragraph: ", text)
        return self._get_block(text).nest("paragraph")
        
    def table(self, header, body):
        #print("table: ", header, body)
        return None
        
    def table_row(self, content):
        #print("table_row", content)
        return None
    
    def table_cell(self, content, **flags):
        #print("table_cell: ", content, flags)
        return None
    
    # ==== Inline elements ====
    
    def autolink(self, link, is_email=False):
        #print("autolink: ", link, is_email)
        return link.inline.apply_marker("l")
    
    def codespan(self, text):
        #print("codespan: ", text)
        return InlineCollector(text).apply_marker("c")
        
    def double_emphasis(self, text):
        return text.inline.apply_marker("i").apply_marker("b")
    
    def emphasis(self, text):
        #print("emphasis: ", text)
        return text.inline.apply_marker("i")
    
    def image(self, src, title, alt_text):
        return None
    
    def linebreak(self):
        return InlineCollector("\n")
    
    def newline(self):
        return None
    
    def link(self, link, title, content):
        #print("link: ", link, title, content)
        self.links.add(link)
        idx = self.links.index(link) + 1
        return content.inline.apply_marker("l").append("[{:d}]".format(idx))
    
    def strikethrough(self, text):
        #print("strikethrough: ", text)
        return text.inline.apply_marker("s")
    
    def text(self, text):
        #print("text: ", text)
        return InlineCollector(text)
    
    def inline_html(self, text):
        return text.inline
    
    def terminate(self):
        """
            Called by the collector when the render function is called
        """
        collector = Collector(self)
        if len(self.links) > 0:
            # Blank line
            collector += BlockCollector(InlineCollector(""))
            # Each link in a special cell with format
            # "[<idx>] <link>"
            for idx, link in enumerate(self.links, start=1):
                content = InlineCollector("[{:d}] ".format(idx))
                content += InlineCollector(link).apply_marker("l")
                block = BlockCollector(content).nest("link")
                collector += block
            # Reset links
            self.links = IndexedList()
        return collector

def convert(inpath, outpath):
    from openpyxl import Workbook
    
    renderer = ExcelRenderer()
    markdown = mistune.Markdown(renderer=renderer)
    with open(inpath, "r", encoding="utf-8") as fin:
        result = markdown(fin.read())
    wb = Workbook()
    result.render(wb.active)
    wb.save(outpath)

def cmdline():
    import argparse
    import os.path
    
    parser = argparse.ArgumentParser(
        description='Convert Markdown to Excel.'
    )
    parser.add_argument('input',
        help='input file path (Markdown)'
    )
    parser.add_argument('output',
        nargs="?", default=None,
        help='output file path (Xlsx)'
    )
    args = parser.parse_args()
    
    inpath = os.path.abspath(args.input)
    outpath = args.output
    
    if outpath is None:
        outpath = os.path.splitext(inpath)[0] + ".xlsx"
    
    convert(inpath, outpath)
    
    return 0


if __name__ == '__main__':
    import sys
    sys.exit(cmdline())
